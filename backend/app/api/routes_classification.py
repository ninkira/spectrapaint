import logging
import os
import re
import uuid
from datetime import datetime, timezone
from functools import lru_cache
from pathlib import Path

import numpy as np
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from pydantic import BaseModel
from sqlalchemy.orm import Session
from spectral import io as spyio

from ..analysis.classification.distance_metrics import DistanceMetrics
from ..analysis.classification.distance_registry import METHODS
from ..analysis.classification.reference_registry import (
    get_reference_library_or_404,
    list_reference_libraries,
)
from ..db.database import get_db
from ..db.ids import stable_id
from ..db.models import (
    DerivedDataset,
    HsiCube,
    ProcessingOperation,
    RoiAnnotation,
    SpectralExtraction,
    SpectralLibrary,
)
from ..services.dataset_store import get_dataset_record_or_404
from ..services.dataset_sync import upsert_spectral_library
from ..paths import APP_DATA_DIR


router = APIRouter()
logger = logging.getLogger(__name__)


@router.get("/classification/methods")
def list_methods():
    return {"methods": METHODS}


@router.get("/classification/libraries")
def list_libraries():
    return {"libraries": list_reference_libraries()}


def _safe_lib_stem(name: str) -> str:
    stem = Path(name).stem
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("_")
    return cleaned or "library"


def _write_upload(upload: UploadFile, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    with open(dest, "wb") as f:
        f.write(upload.file.read())


@router.post("/classification/libraries/upload")
def upload_library(
    header: UploadFile = File(...),
    data: UploadFile = File(...),
    name: str | None = Form(None),
    db: Session = Depends(get_db),
):
    """Register a reference spectral library (ENVI .hdr + .sli/.img) so classification can use it.

    Saved under old_man/spectral_libraries/<name>/ where list_reference_libraries() scans, so it
    becomes selectable immediately. The DB SpectralLibrary row is upserted best-effort.
    """
    if not (header.filename or "").lower().endswith(".hdr"):
        raise HTTPException(status_code=400, detail="The header must be an ENVI .hdr file")
    data_ext = Path(data.filename or "").suffix.lower()
    # The library scanner only looks for a sibling .sli or .img, so normalise to one of those.
    out_ext = ".img" if data_ext == ".img" else ".sli"

    stem = _safe_lib_stem(name or header.filename or "library")
    folder = APP_DATA_DIR / "old_man" / "spectral_libraries" / stem
    hdr_path = folder / f"{stem}.hdr"
    data_path = folder / f"{stem}{out_ext}"

    _write_upload(header, hdr_path)
    _write_upload(data, data_path)

    try:
        spyio.envi.open(str(hdr_path), str(data_path))
    except Exception as exc:
        hdr_path.unlink(missing_ok=True)
        data_path.unlink(missing_ok=True)
        try:
            folder.rmdir()
        except OSError:
            pass
        raise HTTPException(status_code=400, detail=f"Could not read the spectral library: {exc}") from exc

    # Look the library back up so its id/label match exactly what the scanner produces.
    lib = next(
        (item for item in list_reference_libraries() if Path(item["hdr_path"]) == hdr_path),
        None,
    )
    if lib is None:
        raise HTTPException(status_code=500, detail="Library saved but could not be registered")

    try:
        upsert_spectral_library(db, lib)
        db.commit()
    except Exception:
        db.rollback()
        logger.exception("Failed to upsert SpectralLibrary row (library still usable from disk)")

    return {"id": lib["id"], "label": lib["label"]}


class MeanSignal(BaseModel):
    wavelengths_nm: list[float]
    values: list[float]


class PipelineRequest(BaseModel):
    dataset_id: str
    roi_id: str
    preprocessing_method_id: str | None = None
    classification_method_id: str
    reference_library_id: str
    mean_signal: MeanSignal
    top_k: int = 5


def _strip_spectrum_suffix(name: str) -> str:
    """
    Normalize IDs like EB_1_2_p1_sh2 -> EB_1_2_p1 for metadata matching/display.
    """
    return re.sub(r"_sh\d+$", "", str(name).strip(), flags=re.IGNORECASE)


def _norm_token(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def _norm_header(value: object) -> str:
    return _norm_token(str(value))


def _find_name_columns(columns: list[object]) -> tuple[int | None, int | None]:
    prefix_idx = None
    pname_idx = None
    for i, col in enumerate(columns):
        key = _norm_header(col)
        if key == "spectranameprefix":
            prefix_idx = i
        if key in {"pigmentnamepname", "pigmentname"}:
            pname_idx = i
    return prefix_idx, pname_idx


def _rows_from_excel(candidate: Path) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []

    # First try pandas (handles many edge cases cleanly when available).
    try:
        import pandas as pd  # type: ignore

        engines: list[str | None]
        if candidate.suffix.lower() == ".xls":
            engines = ["xlrd", None]
        else:
            engines = ["openpyxl", None]

        for engine in engines:
            try:
                df = pd.read_excel(str(candidate), engine=engine) if engine else pd.read_excel(str(candidate))
            except Exception:
                continue

            cols = [str(c).strip() for c in df.columns]
            prefix_idx, pname_idx = _find_name_columns(cols)
            if prefix_idx is None or pname_idx is None:
                continue

            prefix_col = cols[prefix_idx]
            pname_col = cols[pname_idx]
            for _, rec in df[[prefix_col, pname_col]].iterrows():
                prefix_val = str(rec[prefix_col]).strip()
                pname_val = str(rec[pname_col]).strip()
                if prefix_val:
                    rows.append((prefix_val, pname_val))
            if rows:
                return rows
    except Exception:
        pass

    # Pandas unavailable/failed: fallback to direct readers.
    suffix = candidate.suffix.lower()
    if suffix == ".xls":
        try:
            import xlrd  # type: ignore

            wb = xlrd.open_workbook(str(candidate))
            sh = wb.sheet_by_index(0)
            if sh.nrows <= 1:
                return rows
            headers = [sh.cell_value(0, c) for c in range(sh.ncols)]
            prefix_idx, pname_idx = _find_name_columns(headers)
            if prefix_idx is None or pname_idx is None:
                return rows
            for r in range(1, sh.nrows):
                prefix_val = str(sh.cell_value(r, prefix_idx)).strip()
                pname_val = str(sh.cell_value(r, pname_idx)).strip()
                if prefix_val:
                    rows.append((prefix_val, pname_val))
        except Exception:
            return rows
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook  # type: ignore

            wb = load_workbook(filename=str(candidate), read_only=True, data_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            header_row = next(it, None)
            if not header_row:
                return rows
            prefix_idx, pname_idx = _find_name_columns(list(header_row))
            if prefix_idx is None or pname_idx is None:
                return rows
            for row in it:
                prefix_val = str(row[prefix_idx] if prefix_idx < len(row) else "").strip()
                pname_val = str(row[pname_idx] if pname_idx < len(row) else "").strip()
                if prefix_val:
                    rows.append((prefix_val, pname_val))
        except Exception:
            return rows
    return rows


@lru_cache(maxsize=16)
def _rows_from_excel_cached(path: str, mtime_ns: int) -> tuple[tuple[str, str], ...]:
    """Cached wrapper — returns immutable tuple so lru_cache works."""
    del mtime_ns  # part of cache key for invalidation
    return tuple(_rows_from_excel(Path(path)))


def _resolve_label_for_match(
    library_dir: str,
    pigment_name: str,
) -> dict:
    """Resolve display label from Pigment name (pname) using spectra prefix match."""
    base = Path(library_dir)
    excel_candidates: list[Path] = []
    for p in (
        base / "__pigmentlistZenodo_custom.xls",
        base / "__pigmentlistZenodo_custom.xlsx",
        base / "__pigmentlistZenodo.xls",
        base / "__pigmentlistZenodo.xlsx",
    ):
        if p.exists():
            excel_candidates.append(p)

    # Project-level fallback for shared metadata file(s).
    shared_root = APP_DATA_DIR / "old_man" / "spectral_libraries"
    if shared_root.exists():
        for pattern in ("**/__pigmentlistZenodo_custom.xls", "**/__pigmentlistZenodo_custom.xlsx", "**/__pigmentlistZenodo.xls", "**/__pigmentlistZenodo.xlsx"):
            for p in sorted(shared_root.glob(pattern)):
                if p not in excel_candidates:
                    excel_candidates.append(p)

    label_name = pigment_name
    key_norm = _norm_token(_strip_spectrum_suffix(str(pigment_name).strip()))

    for candidate in excel_candidates:
        if not candidate.exists():
            continue
        try:
            mtime_ns = os.stat(str(candidate)).st_mtime_ns
        except OSError:
            continue
        for spectra_prefix, pname in _rows_from_excel_cached(str(candidate), mtime_ns):
            prefix_norm = _norm_token(_strip_spectrum_suffix(spectra_prefix))
            if not prefix_norm:
                continue
            if key_norm == prefix_norm or key_norm.startswith(prefix_norm) or prefix_norm.startswith(key_norm):
                clean_pname = str(pname).strip()
                if clean_pname:
                    label_name = clean_pname
                return {"label_name": label_name}

    return {"label_name": label_name}


def _parse_spectra_names(raw: object, count: int) -> list[str]:
    if isinstance(raw, list):
        names = [str(x).strip() for x in raw]
    elif isinstance(raw, str):
        cleaned = raw.strip().strip("{}")
        names = [part.strip() for part in cleaned.split(",") if part.strip()]
    else:
        names = []

    if len(names) != count:
        names = [f"spectrum_{i}" for i in range(count)]
    return names


_library_cache: dict[tuple[str, str], tuple[np.ndarray, list[str], list[float]]] = {}
_library_cache_mtimes: dict[tuple[str, str], float] = {}


def _load_library_matrix(reference_library: dict) -> tuple[np.ndarray, list[str], list[float]]:
    hdr_path = reference_library["hdr_path"]
    data_path = reference_library["data_path"]
    cache_key = (hdr_path, data_path)

    try:
        current_mtime = os.stat(hdr_path).st_mtime
    except OSError:
        current_mtime = 0.0

    if cache_key in _library_cache and _library_cache_mtimes.get(cache_key) == current_mtime:
        return _library_cache[cache_key]

    img = spyio.envi.open(hdr_path, data_path)
    if hasattr(img, "spectra"):
        arr = np.asarray(img.spectra, dtype=float)
    else:
        arr = np.asarray(img.load(), dtype=float)

    if arr.ndim == 3:
        arr = arr[:, 0, :]
    elif arr.ndim != 2:
        raise HTTPException(status_code=500, detail=f"Unsupported library array shape: {arr.shape}")

    md = getattr(img, "metadata", {}) or {}
    if hasattr(img, "names") and getattr(img, "names", None):
        names = [str(n).strip() for n in img.names]
        if len(names) != arr.shape[0]:
            names = _parse_spectra_names(md.get("spectra names"), arr.shape[0])
    else:
        names = _parse_spectra_names(md.get("spectra names"), arr.shape[0])

    raw_wl = md.get("wavelength", [])
    if (not raw_wl) and hasattr(img, "bands") and getattr(img.bands, "centers", None):
        raw_wl = img.bands.centers
    wavelengths: list[float] = []
    for w in raw_wl:
        try:
            wavelengths.append(float(w))
        except (TypeError, ValueError):
            wavelengths = []
            break

    result = (arr, names, wavelengths)
    _library_cache[cache_key] = result
    _library_cache_mtimes[cache_key] = current_mtime
    return result


def _compute_distances(method_id: str, query: np.ndarray, library_matrix: np.ndarray) -> np.ndarray:
    dm = DistanceMetrics()
    query_matrix = np.repeat(query[None, :], library_matrix.shape[0], axis=0)

    if method_id == "sam_matrix":
        distances = dm.matrix_spectral_angle_mapper(query_matrix, library_matrix)
    elif method_id == "cosine_matrix":
        distances = dm.matrix_cosine_distance(query_matrix, library_matrix)
    elif method_id == "klpd":
        distances = np.asarray(dm.klpd_spectral(query_matrix, library_matrix, mode=3)).reshape(-1)
    elif method_id == "sam_pixel":
        distances = np.asarray(
            [dm.pixel_spectral_angle_mapper(query, row) for row in library_matrix],
            dtype=float,
        )
    else:
        raise HTTPException(status_code=400, detail=f"Unknown classification method: {method_id}")

    finite_mask = np.isfinite(distances)
    if not np.all(finite_mask):
        distances = np.where(finite_mask, distances, np.inf)
    if distances.shape[0] != library_matrix.shape[0]:
        raise HTTPException(
            status_code=500,
            detail=(
                f"Distance size mismatch for method '{method_id}': "
                f"expected {library_matrix.shape[0]}, got {distances.shape[0]}"
            ),
        )
    return distances


def _persist_classification_run(
    db: Session, req: "PipelineRequest", reference_library: dict,
    results: dict, top_matches: list[dict],
) -> None:
    """Record a run as provenance: SpectralExtraction -> ProcessingOperation -> DerivedDataset."""
    now = datetime.now(timezone.utc)

    # Ensure the referenced SpectralLibrary exists (populated by the startup sync; upsert if not).
    lib_uuid = stable_id("library", req.reference_library_id)
    if db.get(SpectralLibrary, lib_uuid) is None:
        upsert_spectral_library(db, reference_library, now)
        db.flush()

    # SpectralExtraction: one per ROI (upsert). roi_id FK only if that annotation is saved.
    ext_uuid = stable_id("extraction", req.roi_id)
    roi_fk = None
    try:
        candidate = uuid.UUID(str(req.roi_id))
        if db.get(RoiAnnotation, candidate) is not None:
            roi_fk = candidate
    except (ValueError, TypeError):
        pass
    wl = req.mean_signal.wavelengths_nm
    wrange = f"{min(wl):.1f}-{max(wl):.1f} nm" if wl else None
    existing = db.get(SpectralExtraction, ext_uuid)
    if existing is None:
        # No extraction on record — the ROI was never saved, so the client's mean signal is all
        # we have. std/pixel_count are genuinely unknown here rather than zero.
        db.add(SpectralExtraction(
            extraction_id=ext_uuid,
            roi_id=roi_fk,
            library_id=lib_uuid,
            mean_spectrum=list(req.mean_signal.values),
            std_spectrum=[],
            pixel_count=0,
            wavelength_range=wrange,
            extracted_at=now,
        ))
    else:
        # Saving the ROI already extracted real statistics. Record which library this run
        # compared against, but never overwrite mean/std/pixel_count with the client's figures.
        existing.library_id = lib_uuid
    db.flush()  # persist the extraction before the operation references it

    # ProcessingOperation: one row per run (accumulates a provenance history).
    op_uuid = uuid.uuid4()
    db.add(ProcessingOperation(
        operation_id=op_uuid,
        operation_type="classification",
        method_name=req.classification_method_id,
        parameters={
            "preprocessing_method_id": req.preprocessing_method_id,
            "reference_library_id": req.reference_library_id,
            "top_k": req.top_k,
        },
        executed_at=now,
        software_version="SpectraPaint",
        input_extraction_id=ext_uuid,
    ))
    db.flush()  # persist the operation before the derived dataset references it

    # DerivedDataset: the result (ranking) stored in `lookup`; no output file.
    cube_uuid = stable_id("cube", req.dataset_id)
    was_from = cube_uuid if db.get(HsiCube, cube_uuid) is not None else ext_uuid
    db.add(DerivedDataset(
        derived_id=uuid.uuid4(),
        type="classification",
        file_format="json",
        data_ref=None,
        lookup=results,
        classes=len(top_matches),
        class_names=[str(m.get("label_name")) for m in top_matches],
        was_derived_from=was_from,
        operation_id=op_uuid,
        created_at=now,
    ))

    db.commit()


@router.post("/classification/pipeline/run")
def run_pipeline(req: PipelineRequest, db: Session = Depends(get_db)):
    get_dataset_record_or_404(req.dataset_id)
    reference_library = get_reference_library_or_404(req.reference_library_id)

    if not req.mean_signal.values:
        raise HTTPException(status_code=400, detail="mean_signal.values is empty")
    if len(req.mean_signal.wavelengths_nm) != len(req.mean_signal.values):
        raise HTTPException(status_code=400, detail="mean_signal wavelengths/values length mismatch")

    known_methods = {m["id"] for m in METHODS}
    if req.classification_method_id not in known_methods:
        raise HTTPException(status_code=400, detail=f"Unknown classification method: {req.classification_method_id}")

    query = np.asarray(req.mean_signal.values, dtype=float)
    library_matrix, pigment_names, library_wavelengths = _load_library_matrix(reference_library)

    min_bands = min(query.shape[0], library_matrix.shape[1])
    if min_bands <= 0:
        raise HTTPException(status_code=400, detail="No overlapping bands for classification")
    if query.shape[0] != min_bands:
        query = query[:min_bands]
    if library_matrix.shape[1] != min_bands:
        library_matrix = library_matrix[:, :min_bands]
    if library_wavelengths and len(library_wavelengths) >= min_bands:
        library_wavelengths = library_wavelengths[:min_bands]

    distances = _compute_distances(req.classification_method_id, query, library_matrix)
    k = max(1, min(int(req.top_k), int(distances.shape[0])))
    order = np.argsort(distances)[:k]
    library_dir = str(Path(reference_library["hdr_path"]).parent)

    top_matches = []
    for rank, idx in enumerate(order):
        raw_name = pigment_names[int(idx)]
        meta = _resolve_label_for_match(
            library_dir=library_dir,
            pigment_name=raw_name,
        )
        top_matches.append(
            {
                "rank": rank + 1,
                "index": int(idx),
                "pigment_name": raw_name,
                "spectra_name_prefix": _strip_spectrum_suffix(raw_name),
                "label_name": meta.get("label_name", raw_name),
                "score": float(distances[int(idx)]),
                "values": library_matrix[int(idx)].tolist(),
            }
        )

    results = {
        "preprocessing_method": req.preprocessing_method_id,
        "classification_method": req.classification_method_id,
        "reference_library_id": req.reference_library_id,
        "reference_library_label": reference_library["label"],
        "top_matches": top_matches,
    }

    # Record the run as provenance (best-effort — never block the classification response).
    try:
        _persist_classification_run(db, req, reference_library, results, top_matches)
    except Exception:
        db.rollback()
        logger.exception("Failed to persist classification run (results still returned)")

    return {
        "datasetId": req.dataset_id,
        "roiId": req.roi_id,
        "mean_signal": req.mean_signal.model_dump(),
        "library": {
            "bands": int(library_matrix.shape[1]),
            "size": int(library_matrix.shape[0]),
            "wavelengths_nm": library_wavelengths,
        },
        "results": results,
    }
